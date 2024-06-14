#import necessary packages
import streamlit as st
import pandas as pd
import requests
import openpyxl
import os
import io

def vin_data(file_bytes, original_filename):
    #some excel files have more than 1 sheet, we handle excel files with more than 1 sheet by telling the 
    #code to read the sheet named 'Vehicle & Asset List' as this is the standard naming convention
    #write the information from this sheet into dataframe named 'raw_vin_data'
    wb = openpyxl.load_workbook(file_bytes)
    res = len(wb.sheetnames)
    if res > 1:
        raw_vin_data = pd.read_excel(file_bytes, 'Vehicle & Asset List', header=3)
    else:
        raw_vin_data = pd.read_excel(file_bytes, header=3)
    
    #assign new column names to raw_vin_data dataframe for dataframe to standardize raw_vin_data for query
    for column in raw_vin_data.columns:
        if 'vin' in column.lower():
            raw_vin_data.rename(columns={column:'VIN'}, inplace=True)
    
    #create base url that will be augmented with VIN for query
    base_url = 'https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVin/'
    
    #create a dataframe with relevant columns where detailed vehicle information will be stored
    vin_data = pd.DataFrame({'VIN':[], 'VIN Mask': [], 'Model Year': [], 'Manufacturer': [], 
                             'Make': [], 'Model': [], 'Trim': [], 'Weight Class': [],
                             'Body/Cab Type': [], 'Body Class': [], 'Drive Type': [],
                             'Fuel Type': [], 'Engine Model': [], 'Engine Configuration': [],
                             'Engine Cyl': [], 'Displacement (Litres)': [], 'Engine Horse Power': [],
                             'Transmission': [], 'Speeds': [], 'Error Test': []
                            })
    
    #select VINs from original document, exclude empty cells (NAN/NULL values)
    values = [raw_vin_data['VIN'][i] for i in raw_vin_data.index  if pd.isna(raw_vin_data['VIN'][i]) == False]
    
    #create list to store dictionaries, each dictionary will relate to a specific VIN or row of the dataframe
    results = []
    
    #create variable to keep track of which index is being used, this keeps track of what row of the dataframe
    #the code is on
    ind = 0
    
    #query the NHTSA VIN database using each VIN from the original sales document to collect info on vehicle 
    #year, make, model, fuel, and vehicle type, as MCF operates in United States all entries for Country = US
    
    #iterate through each VIN in list of VINs
    for value in values:
        #ensure the type of the VIN is string
        value = str(value)
        #remove spaces from VIN, accounts for common data entry error
        value = value.replace(" ", "")
        #create VIN specific link to access details for API query
        url = base_url + value + '?format=json'
        #pulls details from url, bypasses certification verification error created by Michelin firewalls
        response = requests.get(url, verify = False)
        #check to see if vin is accurate, if accurate extract data into dictionary and add to results list
        try:
            #save url information as data variable for query
            data = response.json()
            #create key for decoding desired information from url data
            decoded_values = {item['Variable']: item['Value'] for item in data['Results']}
            #create a dictionary with vehicle information from VIN query, information based on specific VIN
            results.append({
                'VIN': value, 
                'VIN Mask': decoded_values.get('Vehicle Descriptor', 'N/A'), 
                'Model Year': decoded_values.get('Model Year', 'N/A'), 
                'Manufacturer': decoded_values.get('Manufacturer Name', 'N/A'), 
                'Make': decoded_values.get('Make', 'N/A'), 
                'Model': decoded_values.get('Model', 'N/A'), 
                'Trim': decoded_values.get('Trim', 'N/A'), 
                'Weight Class': decoded_values.get('Gross Vehicle Weight Rating From', 'N/A'),
                'Body/Cab Type': decoded_values.get('Cab Type', 'N/A'), 
                'Body Class': decoded_values.get('Body Class', 'N/A'), 
                'Drive Type': decoded_values.get('Drive Type', 'N/A'),
                'Fuel Type': decoded_values.get('Fuel Type - Primary', 'N/A'), 
                'Engine Model': decoded_values.get('Engine Model', 'N/A'), 
                'Engine Configuration': decoded_values.get('Engine Configuration', 'N/A'),
                'Engine Cyl': decoded_values.get('Engine Number of Cylinders', 'N/A'), 
                'Displacement (Litres)': decoded_values.get('Displacement (L)', 'N/A'), 
                'Engine Horse Power': decoded_values.get('Engine Brake (hp) From', 'N/A'),
                'Transmission': decoded_values.get('Transmission Style', 'N/A'), 
                'Speeds': decoded_values.get('Transmission Speeds', 'N/A'), 
                'Error Test': decoded_values.get('Error Text', 'N/A')
            })
            #increase the index by 1, indicates code moves onto next VIN/row
            ind += 1
        #if vin not accurate, use error handling, will only move to this step if url produces empty response 
        #(data variable is empty)
        except json.JSONDecodeError as e:
            results.append({
                'VIN': value, 
                'VIN Mask': 'Error', 
                'Model Year': 'Error', 
                'Manufacturer': 'Error', 
                'Make': 'Error', 
                'Model': 'Error', 
                'Trim': 'Error', 
                'Weight Class': 'Error',
                'Body/Cab Type': 'Error', 
                'Body Class': 'Error', 
                'Drive Type': 'Error',
                'Fuel Type': 'Error', 
                'Engine Model': 'Error', 
                'Engine Configuration': 'Error',
                'Engine Cyl': 'Error', 
                'Displacement (Litres)': 'Error', 
                'Engine Horse Power': 'Error',
                'Transmission': 'Error', 
                'Speeds': 'Error', 
                'Error Test': 'Error: Incorrect VIN, no data exists'
            })
            #increase the index by 1, indicates code moves onto next VIN/row
            ind += 1
        #if code times out, this error handling will make sure the code does not run indefinitely, if
        #encountered the code will stop processing VINs and communicate a time out error to the user
        except requests.exceptions.Timeout as e:
            return "Timed out"
        
    #create dataframe from list of dictionaries, each dictionary is a row within the 'results' dataframe
    results = pd.DataFrame(results)
    
    #remove duplicate VINs
    results.drop_duplicates(subset=['VIN'],inplace= True)
    
    #information should be written to an Excel file, the output file will have the same name as the input
    #file with _VIN_data appended
    file_path = os.path.splitext(file_path)[0] + "_VIN_data.xlsx"
    pd.DataFrame(results).to_excel(file_path, index=False)
    
    #write dataframe to Excel file table
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        
        #create an Excel sheet named 'Vehicle Data' to hold the dataframe
        results.to_excel(writer, index=False, sheet_name='Vehicle Data')

        #access Excel file and worksheet 
        workbook = writer.book
        worksheet = writer.sheets['Vehicle Data']

        #iterate through columns, find the max width of the cells in the column
        for idx, column in enumerate(worksheet.columns):
            #skip over 'ERROR CODE' column as it is the last column, formatting is unneccesary here
            if worksheet.cell(row=1, column=idx + 1).value != 'Error Test':
                max_length = 0
                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                    #adjust column width to show all data
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width
            #adjust 'ERROR CODE' column to be the width of the title
            if worksheet.cell(row=1, column=idx + 1).value == 'Error Test':
                worksheet.column_dimensions[chr(65 + idx)].width = 12
    #ensure curser is at the beginning of the buffer before reading or downloading information
    buffer.seek(0)
    #save and return the processed excel file name and the buffer (memory of object)
    return buffer, processed_filename

#set the text font as open sans to adhere to Michelin branding guidelines
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css?family=Your+Font+Name');
body {
    font-family: 'Your Font Name', open-sans;
}
</style>
""", unsafe_allow_html=True)

#add the Michelin banner to the top of the application, if the image link breaks you can correct this by copying and
#pasting an alternative image url in the ()
st.image("https://www.tdtyres.com/wp-content/uploads/2018/12/kisspng-car-michelin-man-tire-logo-michelin-logo-5b4c286206fa03.5353854915317177300286.png")

#set the application title to 'VIN Vehicle Data'
st.title("VIN Vehicle Data")

#create a drag and drop box for file uploading, indicate that the file must be a CSV or Excel file
uploaded_file = st.file_uploader("Upload an Excel file", type=["xls", "xlsx", "csv"])

#if a file is uploaded call vin_data function to process the uploaded file
if uploaded_file:
    file_bytes = uploaded_file.read()
    buffer, processed_filename = vin_data(io.BytesIO(file_bytes), uploaded_file.name)
    #if the file is successfully processed inform the user
    st.success(f'File "{uploaded_file.name}" successfully processed.')
    #create button to download processed output file
    st.download_button(label="Download Processed File", data=buffer, file_name=processed_filename)

#document how to use the "VIN Vehicle Data application to the user

st.markdown('''This application checks customer VINs with the [National Highway Traffic Safety Administration API](https://vpic.nhtsa.dot.gov/api/) to retrieve vehicle information based on the VIN. This application can handle large volumes of VINs but greater numbers of uploaded VINs will slow down processing time. Processing 2200 VINs takes roughly 25 minutes. When uploading large numbers of VINs please be patient and do not close out the application while processing

**Input Document Requirements:**

- The uploaded document containing the VINs must follow the standard [Michelin Connected Fleet Deployment Template.](https://view.officeapps.live.com/op/view.aspx?src=https%3A%2F%2Fraw.githubusercontent.com%2FChanMichelin%2FautovinMCF%2Fmain%2Fexamples%2FMCF%2520Deployment%2520Template.xlsx&wdOrigin=BROWSELINK) This application cannot decipher different document formats. If an error is indicated with a file you upload, please check the uploaded document follows the formatting guidelines.
- The VIN column must include the VINs the user wants to query. This is the only field necessary to retrieve vehicle data. 

***Example Input File:*** [***VIN Example***](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7BFA264B31-B424-418C-8D1C-C0E5F001094E%7D&file=MCF%20Deployment%20Template.xlsx&action=default&mobileredirect=true&wdsle=0)

***Note:*** If you are interested in vehicle information regarding VINs recorded in a different format/document download the MCF Deployment Template linked above, then copy and paste the VINs into the VIN column and upload this document for bulk processing.

**Output Document Description:**

- This application processes all the VINs regardless of VIN accuracy or vehicle type. 
- If the VIN is inaccurate or relates to a lift/trailer not present in the NHTSA database the 'Error' column will indicate what type of error is occurring for user reference. 
- An error code of 0 indicates there was no issue with the VIN. 
- This file provides information on vehicle make, model, year, and manufacturer as well as more detailed information pertaining to trim, engine type, primary fuel etc. 
- When a cell is empty, but the error column indicates there was no issue processing the VIN (error code is 0) this indicates that data on this vehicle specification is not recorded within the NHTSA database. 
- The output Excel file will have the same name as the original document followed by _VIN_data. 

***Example Output File:*** [***VIN Example_VIN_data***](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7B7481464E-023E-4E40-9007-34AE4022EECE%7D&file=VIN%20Example_VIN_data.xlsx&action=default&mobileredirect=true&wdsle=0)

If you are interested in a list of accurate VINs that relate to CAN compatible vehicles excluding trailers and lifts, please refer to the [Automated VIN Decoding Application.](https://autovin.streamlit.app/)

If you are encountering issues with this application please contact Massaer Diouf.
''')